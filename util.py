import openpyxl
import os

fl=os.listdir('./excel')

def excel(name):
    all_values = []
    for i in fl:
        if name in i:
            workbook = openpyxl.load_workbook('./excel/'+i,data_only=True)

            worksheet=workbook['목록']

            for row in worksheet.rows:
                row_value = []
                for cell in row:
                    row_value.append(cell.value)
                all_values.append(row_value)
    return all_values

def excel():
    all_values = []
    for i in fl:
        print(str(i)+' 엑셀파일 가져오는 중')
        workbook = openpyxl.load_workbook('./excel/'+i,data_only=True)

        worksheet=workbook['목록']
        count = 1
        for row in worksheet.rows:
            row_value = []
            worksheet.cell(row=count, column=55).value=''
            count=count+1
            for cell in row:
                row_value.append(cell.value)
            all_values.append(row_value)
    workbook.save('./excel/'+i)
    return all_values

def saveExcel(values,reason):
    for i in fl:
        workbook = openpyxl.load_workbook('./excel/'+i,data_only=True)

        worksheet=workbook['목록']
        for j, k in zip(values, reason):
            worksheet.cell(row=int(j), column=55).value=k

        workbook.save('./excel/'+i)


def findImage(name):
    fullName=''
    fl=os.listdir('./image')
    for i in fl:
        if name in i:
            fullName = os.getcwd()+'/image/'+i

    return str(fullName)


def findImage(name,list):
    fullName=''
    fl=os.listdir('./image/'+list)
    for i in fl:
        if name in i:
            fullName = os.getcwd()+'/image/'+list+'/'+i

    return str(fullName)
